from selenium import webdriver

import openpyxl



#---------DECLARING VARIABLES -------------

ExcelPath = "//home//elyor//Selenium//EXCEL.xlsx"

chromePath = "/home/elyor/Selenium/chromedriver"

geckoPath = "/home/elyor/Selenium/geckodriver"

URL = "https://www.surveymonkey.com/"


#---------END OF THE DECLARING VARIABLES -------------


#---DECLARING WEBDRIVER CHROME or IE...---
driver = webdriver.Chrome(executable_path= chromePath)


#--- DELETE ALL THE COOKIES BEFORE START ---
driver.delete_all_cookies()

#---DECLARE IMPLICIT WAIT FOR ALL OBJECT---
driver.implicitly_wait(5)

#---MAXIMIZE THE WINDOW ---
driver.maximize_window()


#----------------HOW TO WRITE THE DATA TO EXCEL FILE --------


workbook = openpyxl.load_workbook(ExcelPath)

sheet = workbook.active         #ALTERNATE: sheet = workbook.get_sheet_by_name("Sheet1")


#1 Here Finding Number of ROWS

RowCount = sheet.max_row

print(RowCount)



#2 Here Column Count

ColumnCount = sheet.max_column

print(ColumnCount)



#3 To Write Value in the Table we'll use 2 Different FOR LOOP


for r in range(1, RowCount+1):

    for c in range(1, ColumnCount+1):

        #Here we'll set the value in Excel

        sheet.cell(row=r, column=c).value = "Hello World"

workbook.save(ExcelPath)

#------------------------------------





#1 Launching URL
#driver.get(URL)
