import openpyxl

#1 Row Count
def getRowCount(file, sheetName):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook.get_sheet_by_name(sheetName)

    return(sheet.max_row)


#2 Column Count
def getColumnCount(file, sheetName):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook.get_sheet_by_name(sheetName)

    return (sheet.max_Column)


#3 Read DATA
def readData(file, sheetName, rownum, columnno):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook.get_sheet_by_name(sheetName)

    return sheet.cell(row=rownum, column=columnno).value


#4 Write DATA
def writeData(file, sheetName, rownum, columnno, data):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook.get_sheet_by_name(sheetName)

    sheet.cell(row=rownum, column=columnno).value = data

    workbook.save(file)
    





















